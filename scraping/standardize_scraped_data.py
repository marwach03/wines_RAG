"""
╔══════════════════════════════════════════════════════════════════╗
║  SCRAPED WINE DATA — USEFULNESS CHECK + STANDARDIZATION          ║
║  Produces 3 pairs of reports: raw-only, raw_2-only, and combined ║
╚══════════════════════════════════════════════════════════════════╝
"""

from __future__ import annotations

import argparse
import glob
import json
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path
from typing import Optional

from langdetect import detect, LangDetectException
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── CONFIG ───────────────────────────────────────────────────────────────────
SCRIPT_DIR          = Path(__file__).parent
DEFAULT_INPUT_DIR   = SCRIPT_DIR / "data" / "raw"
DEFAULT_INPUT_DIR_2 = SCRIPT_DIR / "data" / "raw_2"
DEFAULT_OUTPUT_DIR  = SCRIPT_DIR / "output"
DEFAULT_MIN_LENGTH  = 100

KNOWN_GHOST_TEXTS = {
    "Suchen",
    "Herausforderung Fassweinmarkt – regionale Unterschiede Christian Bauer 27. Februar 2026",
}

DWV_LOGIN_WALL_MARKER = "Sie haben keine Berechtigung"
INDEX_PAGE_URL_MARKERS = ("/tag/", "/category/", "/categories/")

REDDIT_DENSITY_MIN_LENGTH = 500
REDDIT_DENSITY_THRESHOLD  = 1.0

SOURCE_DEFAULT_COUNTRY: dict[str, str] = {
    "weinforum":             "de",
    "deutsches_weininstitut": "de",
}


# ─── SHARED HELPERS ───────────────────────────────────────────────────────────

def normalize(text: str) -> str:
    nfkd = unicodedata.normalize("NFKD", text)
    return "".join(c for c in nfkd if not unicodedata.combining(c)).lower().strip()


def slug(text: str) -> str:
    s = normalize(text)
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return s.strip("_")[:60]


def detect_language(text: str) -> Optional[str]:
    try:
        return detect(text)
    except LangDetectException:
        return None


# ─── TEXT CLEANERS ────────────────────────────────────────────────────────────

def strip_deutsches_nav(text: str) -> str:
    for marker in ("www.deutscheweine.de ", "LinkedIn Facebook "):
        idx = text.find(marker)
        if idx >= 0:
            return text[idx + len(marker):]
    if text.startswith("Startseite "):
        return text[len("Startseite "):]
    return text


def strip_winefolly_nav(text: str) -> str:
    idx = text.find("Written by")
    if idx > 0:
        return text[idx:]
    return text


# ─── PER-SOURCE EXTRACTORS ────────────────────────────────────────────────────

def extract_decanter(entry: dict) -> Optional[dict]:
    meta = entry.get("metadata", {})
    return {
        "source":    "decanter",
        "url":        meta.get("url"),
        "title":       meta.get("title"),
        "text":         entry.get("text", ""),
        "post_type":      entry.get("post_type"),
        "extra": {
            "drinking_window": meta.get("drinking_window"),
            "timestamp":        meta.get("timestamp"),
        },
    }


def extract_reddit(entry: dict) -> Optional[dict]:
    return {
        "source":    "reddit",
        "url":        entry.get("url"),
        "title":       entry.get("raw_title"),
        "text":         entry.get("text", ""),
        "post_type":      entry.get("post_type"),
        "extra": {
            "author":              entry.get("author"),
            "score":                entry.get("score"),
            "subreddit":             entry.get("subreddit"),
            "created_at":             entry.get("created_at"),
            "keywords_matched":        entry.get("metadata", {}).get("keywords_matched", []),
            "countries_detected":       entry.get("metadata", {}).get("countries_detected", []),
        },
    }


def extract_old(entry: dict) -> Optional[dict]:
    result = extract_reddit(entry)
    if result is not None:
        result["source"] = "old"
    return result


def extract_vinepair(entry: dict) -> Optional[dict]:
    meta = entry.get("metadata", {})
    return {
        "source":    "vinepair",
        "url":        meta.get("url"),
        "title":       entry.get("title"),
        "text":         entry.get("text", ""),
        "post_type":      entry.get("category"),
        "extra": {
            "scraped_at": meta.get("scraped_at"),
        },
    }


def extract_deutsches_weininstitut(entry: dict) -> Optional[dict]:
    meta = entry.get("metadata", {})
    sd = meta.get("structured_data", {}) if isinstance(meta.get("structured_data"), dict) else {}
    raw_text = entry.get("text", "")
    return {
        "source":    "deutsches_weininstitut",
        "url":        meta.get("url"),
        "title":       sd.get("title") or meta.get("title"),
        "text":         strip_deutsches_nav(raw_text),
        "post_type":      entry.get("post_type"),
        "extra": {
            "keywords_matched": meta.get("keywords_matched", []),
        },
    }


def extract_weinforum(entry: dict) -> Optional[dict]:
    meta = entry.get("metadata", {})
    sd = meta.get("structured_data", {}) if isinstance(meta.get("structured_data"), dict) else {}

    raw_posts = sd.get("posts") or []
    if raw_posts:
        seen: set[str] = set()
        unique_posts = []
        for p in raw_posts:
            p_stripped = p.strip()
            if p_stripped and p_stripped not in seen:
                seen.add(p_stripped)
                unique_posts.append(p_stripped)
        text = "\n\n".join(unique_posts)
    else:
        text = entry.get("text", "")

    return {
        "source":    "weinforum",
        "url":        sd.get("thread_url") or meta.get("url"),
        "title":       sd.get("thread_title"),
        "text":         text,
        "post_type":      entry.get("post_type"),
        "extra": {
            "category":          sd.get("category") or meta.get("category"),
            "post_count":         sd.get("post_count"),
            "wines_mentioned":     sd.get("wines_mentioned", []),
            "regions_mentioned":    sd.get("regions_mentioned", []),
            "topics_discussed":      sd.get("topics_discussed", []),
        },
    }


def extract_winefolly(entry: dict) -> Optional[dict]:
    meta = entry.get("metadata", {})
    sd = meta.get("structured_data", {}) if isinstance(meta.get("structured_data"), dict) else {}
    raw_text = entry.get("text", "")
    grape = meta.get("grape", "")
    title = grape.replace("-", " ").title() if grape else None

    return {
        "source":    "winefolly",
        "url":        meta.get("url"),
        "title":       title,
        "text":         strip_winefolly_nav(raw_text),
        "post_type":      entry.get("post_type"),
        "extra": {
            "grape":            grape,
            "characteristics":   sd.get("characteristics", {}),
            "flavors":            list(dict.fromkeys(sd.get("flavors", []))),
            "keywords_matched":    meta.get("keywords_matched", []),
        },
    }


EXTRACTORS = {
    "decanter":   extract_decanter,
    "reddit":     extract_reddit,
    "old":        extract_old,
    "vinepair":   extract_vinepair,
    "deutsches":  extract_deutsches_weininstitut,
    "weinforum":  extract_weinforum,
    "winefolly":  extract_winefolly,
}

KNOWN_BROKEN_FOLDERS = {
    "jancis_robinson",
    "merum",
    "weinmonitor",
    "vdp",
    "dwv",
    "vinum",
    "vinous",
}

FOLDERS_NEEDING_DEDICATED_EXTRACTOR = {
    "falstaff", "meininger", "winesearcher",
}


def get_source_folder(path: Path, input_root: Path) -> str:
    try:
        rel = path.relative_to(input_root)
        return rel.parts[0]
    except ValueError:
        return path.parent.name


def get_country_from_path(path: Path, input_root: Path) -> Optional[str]:
    try:
        rel = path.relative_to(input_root)
    except ValueError:
        return None
    parts = rel.parts
    if len(parts) >= 3 and parts[0] in {"old", "reddit"}:
        return parts[1]
    return None


def extract_generic(entry: dict) -> Optional[dict]:
    meta = entry.get("metadata", {}) if isinstance(entry.get("metadata"), dict) else {}
    sd = meta.get("structured_data", {}) if isinstance(meta.get("structured_data"), dict) else {}

    text = (
        entry.get("text") or entry.get("content") or entry.get("body")
        or sd.get("full_text") or sd.get("content") or ""
    )
    url = (
        entry.get("url") or entry.get("link")
        or meta.get("url") or sd.get("url")
    )
    title = (
        entry.get("title") or entry.get("raw_title") or entry.get("headline")
        or meta.get("title") or sd.get("title")
    )
    post_type = entry.get("post_type") or entry.get("category") or meta.get("data_type")

    if not text and not url:
        return None

    return {
        "source":    None,
        "url":        url,
        "title":       title,
        "text":         text,
        "post_type":      post_type,
        "extra":           {"raw_top_level_keys": list(entry.keys())},
        "_used_generic_extractor": True,
    }


# ─── QUALITY FILTERS ──────────────────────────────────────────────────────────

def is_ghost_text(text: str) -> bool:
    return text.strip() in KNOWN_GHOST_TEXTS


def is_login_wall(text: str) -> bool:
    return DWV_LOGIN_WALL_MARKER in text


def is_index_page_url(url: Optional[str]) -> bool:
    if not url:
        return False
    return any(marker in url for marker in INDEX_PAGE_URL_MARKERS)


def is_wine_relevant_reddit(text: str) -> bool:
    if len(text) <= REDDIT_DENSITY_MIN_LENGTH:
        return True
    count = text.lower().count("wine") + text.lower().count("wein")
    density = count / (len(text) / 1000)
    return density >= REDDIT_DENSITY_THRESHOLD


def passes_quality_filters(extracted: dict, min_length: int) -> tuple[bool, Optional[str]]:
    text = (extracted.get("text") or "").strip()
    url = extracted.get("url")
    source = extracted.get("source") or extracted.get("_source_folder", "")

    if not text:
        return False, "empty_text"
    if is_ghost_text(text):
        return False, "ghost_text"
    if is_login_wall(text):
        return False, "login_wall"
    if is_index_page_url(url):
        return False, "index_page_url"
    if len(text) < min_length:
        return False, "too_short"
    if not url:
        return False, "missing_url"
    if source in {"reddit", "old"} and not is_wine_relevant_reddit(text):
        return False, "off_topic_low_density"

    return True, None


# ─── LOADING + EXTRACTION ─────────────────────────────────────────────────────

def _is_relative_to(path: Path, root: Path) -> bool:
    try:
        path.relative_to(root)
        return True
    except ValueError:
        return False


def _load_json_or_jsonl(path: Path) -> list[dict]:
    with open(path, "r", encoding="utf-8") as f:
        content = f.read()

    try:
        data = json.loads(content)
    except json.JSONDecodeError:
        entries = []
        for line_no, line in enumerate(content.splitlines(), start=1):
            line = line.strip()
            if not line:
                continue
            entries.append(json.loads(line))
        return entries

    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        return [data]
    raise ValueError(f"unexpected top-level JSON type: {type(data)}")


def infer_folder_from_entry(entry: dict) -> str:
    if entry.get("subreddit") == "winefolly":
        return "winefolly"
    if entry.get("source") == "reddit" or entry.get("subreddit"):
        return "reddit"
    return entry.get("source") or "raw_2_unknown"


def load_raw_files(
    paths: list[Path],
    input_root: Optional[Path],
    input_root_2: Optional[Path] = None,
) -> list[dict]:
    all_entries = []
    for path in paths:
        try:
            data = _load_json_or_jsonl(path)
        except Exception as e:
            print(f"  WARN: could not read {path}: {e}", file=sys.stderr)
            continue

        is_raw2 = input_root_2 is not None and _is_relative_to(path, input_root_2)

        for entry in data:
            entry["_source_file"] = str(path)
            entry["_root"] = "raw_2" if is_raw2 else "raw"
            if is_raw2:
                entry["_source_folder"] = infer_folder_from_entry(entry)
                entry["_country"] = None
            else:
                entry["_source_folder"] = get_source_folder(path, input_root)
                entry["_country"] = get_country_from_path(path, input_root)
            all_entries.append(entry)
    return all_entries


def extract_all(raw_entries: list[dict]) -> tuple[list[dict], Counter, Counter]:
    extracted = []
    broken_counts = Counter()
    generic_counts = Counter()

    for entry in raw_entries:
        folder = entry["_source_folder"]

        if folder in KNOWN_BROKEN_FOLDERS:
            broken_counts[folder] += 1
            continue

        extractor = EXTRACTORS.get(folder)
        if extractor is not None:
            result = extractor(entry)
        else:
            result = extract_generic(entry)
            if result is not None:
                result["source"] = folder
                generic_counts[folder] += 1

        if result is not None:
            result["_source_file"] = entry.get("_source_file")
            result["_source_folder"] = folder
            result["_country"] = entry.get("_country")
            result["_root"] = entry.get("_root")
            result.setdefault("_used_generic_extractor", False)
            extracted.append(result)

    return extracted, broken_counts, generic_counts


# ─── DEDUPLICATION ────────────────────────────────────────────────────────────

def deduplicate_by_url(entries: list[dict]) -> tuple[list[dict], int]:
    seen_urls = set()
    unique = []
    n_duplicates = 0
    for e in entries:
        url = e.get("url")
        key = url or f"__nourl__{id(e)}"
        if key in seen_urls:
            n_duplicates += 1
            continue
        seen_urls.add(key)
        unique.append(e)
    return unique, n_duplicates


# ─── STANDARDIZED CHUNK BUILDING ──────────────────────────────────────────────

def build_chunk(extracted: dict, index: int) -> dict:
    source = extracted["source"]
    text = extracted["text"].strip()
    title = extracted.get("title") or text[:60]

    return {
        "chunk_id":                   f"scraped_{source}_{index:04d}_{slug(title)}",
        "chunk_type":                  extracted.get("post_type") or "article",
        "source":                       source,
        "country":                       extracted.get("_country") or SOURCE_DEFAULT_COUNTRY.get(source),
        "language":                       detect_language(text),
        "title":                           title,
        "text":                             text,
        "url":                               extracted.get("url"),
        "extra":                              extracted.get("extra", {}),
        "_used_generic_extractor":             extracted.get("_used_generic_extractor", False),
        "_source_file":                         extracted.get("_source_file"),
    }


# ─── EXCEL REPORT BUILDING ─────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", start_color="2F5496", end_color="2F5496")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
TITLE_FONT  = Font(bold=True, name="Arial", size=14, color="2F5496")
BODY_FONT   = Font(name="Arial", size=10)
FLAG_FILL   = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
BROKEN_FILL = PatternFill("solid", start_color="F8CBAD", end_color="F8CBAD")
THIN_BORDER = Border(*(Side(style="thin", color="D9D9D9"),) * 4)


def _style_header_row(ws, row: int, n_cols: int):
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _autosize(ws, widths: dict[int, int]):
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def _write_table(ws, start_row: int, headers: list[str], rows: list[list], col_widths: list[int]):
    for j, h in enumerate(headers, start=1):
        ws.cell(row=start_row, column=j, value=h)
    _style_header_row(ws, start_row, len(headers))
    for i, row in enumerate(rows, start=start_row + 1):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            if j == 1:
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="center")
    _autosize(ws, {j: w for j, w in enumerate(col_widths, start=1)})
    return start_row + len(rows) + 1


def build_xlsx_report(report: dict, output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 20

    row = 1

    def _section_title(label: str) -> int:
        nonlocal row
        ws.cell(row=row, column=1, value=label).font = TITLE_FONT
        row += 1
        return row

    _section_title("Wine Data Pipeline — Usefulness Report")

    n_raw   = report["n_raw_entries"]
    n_final = report["n_final_chunks"]
    yield_pct = round(100 * n_final / n_raw, 1) if n_raw else 0

    summary_rows = [
        ["Raw entries loaded", n_raw],
        ["Entries extracted", report["n_extracted"]],
        ["Entries excluded (broken folders)", sum(report["n_excluded_broken_folder"].values())],
        ["Entries passed quality filters", report["n_passed_quality_filters"]],
        ["Entries rejected by quality filters", report["n_rejected_quality_filters"]],
        ["Duplicates removed (by URL)", report["n_duplicates_removed"]],
        ["Final usable chunks", n_final],
        ["Overall usefulness rate", f"{yield_pct}%"],
        ["Chunks built via unverified generic extractor", report["n_final_chunks_via_generic_extractor"]],
    ]
    row = _write_table(ws, row, ["Metric", "Value"], summary_rows, [45, 18])
    row += 1

    ws.cell(row=row, column=1, value="Known broken folders (excluded entirely)").font = Font(bold=True, name="Arial", size=11)
    row += 1
    broken_rows = [[f] for f in report["known_broken_folders"]]
    row = _write_table(ws, row, ["Folder"], broken_rows, [30])
    row += 2

    _section_title("Funnel by Source")

    funnel = report["per_folder_funnel"]
    funnel_rows = []
    for folder, d in sorted(funnel.items()):
        found, extracted, kept = d["found"], d["extracted"], d["kept"]
        rate = round(100 * kept / found, 1) if found else 0
        funnel_rows.append([folder, d.get("root", ""), found, extracted, kept, f"{rate}%", "Broken" if d["broken"] else ""])

    funnel_start = row
    row = _write_table(
        ws, row,
        ["Source folder", "Root", "Found (raw)", "Extracted", "Kept (final)", "Yield %", "Status"],
        funnel_rows, [24, 10, 13, 12, 13, 10, 10],
    )
    for i, frow in enumerate(funnel_rows, start=funnel_start + 1):
        if frow[6] == "Broken":
            for col in range(1, 8):
                ws.cell(row=i, column=col).fill = BROKEN_FILL
    row += 2

    _section_title("Rejection Reasons")

    rej_rows = [[reason, count] for reason, count in sorted(report["rejection_reasons"].items(), key=lambda x: -x[1])]
    row = _write_table(ws, row, ["Rejection reason", "Count"], rej_rows, [28, 12])
    row += 1

    ws.cell(row=row, column=1, value="By folder").font = Font(bold=True, name="Arial", size=11)
    row += 1
    by_folder = report["rejection_reasons_by_folder"]
    all_reasons = sorted(report["rejection_reasons"].keys())
    by_folder_rows = []
    for folder, counts in sorted(by_folder.items()):
        by_folder_rows.append([folder] + [counts.get(r, 0) for r in all_reasons])
    row = _write_table(
        ws, row,
        ["Source folder"] + all_reasons,
        by_folder_rows, [22] + [16] * len(all_reasons),
    )
    row += 2

    _section_title("Final Distribution")

    for dist_title, dist in [
        ("By source", report["final_by_source"]),
        ("By language", report["final_by_language"]),
        ("By country", report["final_by_country"]),
    ]:
        ws.cell(row=row, column=1, value=dist_title).font = Font(bold=True, name="Arial", size=11)
        row += 1
        dist_rows = sorted(dist.items(), key=lambda x: -x[1])
        row = _write_table(ws, row, ["Value", "Count"], [list(r) for r in dist_rows], [26, 12])
        row += 1

    wb.save(output_path)


# ─── MAIN PIPELINE (single run over a given set of inputs) ───────────────────

def run(
    input_paths: list[Path],
    input_root: Optional[Path],
    output_path: Path,
    report_path: Path,
    xlsx_report_path: Path,
    min_length: int,
    input_root_2: Optional[Path] = None,
    label: str = "",
) -> dict:
    tag = f"[{label}] " if label else ""
    print(f"{tag}[1/7] Loading {len(input_paths)} raw file(s)...")
    raw_entries = load_raw_files(input_paths, input_root, input_root_2)
    print(f"  {len(raw_entries)} total raw entries loaded")

    folders_seen = Counter(e["_source_folder"] for e in raw_entries)
    print(f"  Folders found: {dict(folders_seen)}")

    print(f"\n{tag}[2/7] Extracting per-folder schemas...")
    extracted, broken_counts, generic_counts = extract_all(raw_entries)
    print(f"  {len(extracted)} entries extracted")
    if broken_counts:
        print(f"  Excluded (known broken folder): {dict(broken_counts)}")
    if generic_counts:
        print(f"  WARN: used generic fallback extractor for: {dict(generic_counts)}")

    print(f"\n{tag}[3/7] Applying quality filters...")
    passed, rejected = [], []
    rejection_reasons = Counter()
    rejection_reasons_by_folder: dict[str, Counter] = defaultdict(Counter)
    for e in extracted:
        ok, reason = passes_quality_filters(e, min_length)
        if ok:
            passed.append(e)
        else:
            rejected.append({**e, "_rejection_reason": reason})
            rejection_reasons[reason] += 1
            rejection_reasons_by_folder[e["_source_folder"]][reason] += 1
    print(f"  {len(passed)} passed, {len(rejected)} rejected")

    print(f"\n{tag}[4/7] Deduplicating by URL...")
    deduped, n_duplicates = deduplicate_by_url(passed)
    print(f"  {len(deduped)} unique entries kept, {n_duplicates} duplicate(s) removed")

    print(f"\n{tag}[5/7] Building standardized chunks...")
    chunks = [build_chunk(e, i) for i, e in enumerate(deduped, start=1)]

    by_source   = Counter(c["source"] for c in chunks)
    by_language = Counter(c["language"] or "unknown" for c in chunks)
    by_country  = Counter(c["country"] or "n/a" for c in chunks)
    n_generic   = sum(1 for c in chunks if c["_used_generic_extractor"])
    print(f"  By source   : {dict(by_source)}")
    print(f"  By language : {dict(by_language)}")
    print(f"  By country  : {dict(by_country)}")
    print(f"  Built via generic (unverified) extractor: {n_generic}/{len(chunks)}")

    extracted_by_folder = Counter(e["_source_folder"] for e in extracted)
    deduped_by_folder   = Counter(e["_source_folder"] for e in deduped)
    folder_roots: dict[str, set] = defaultdict(set)
    for e in raw_entries:
        folder_roots[e["_source_folder"]].add(e["_root"])
    print("\n  Per-folder funnel (found → extracted → kept after filters+dedup):")
    for folder in sorted(folders_seen):
        found = folders_seen[folder]
        ext   = extracted_by_folder.get(folder, 0)
        kept  = deduped_by_folder.get(folder, 0)
        root  = "+".join(sorted(folder_roots.get(folder, [])))
        flag  = "  ⚠ ZERO SURVIVED" if found > 0 and ext > 0 and kept == 0 else ""
        broken_flag = "  ✗ BROKEN FOLDER (excluded)" if folder in KNOWN_BROKEN_FOLDERS else ""
        print(f"    {folder:<22} root={root:<8} found={found:<6} extracted={ext:<6} kept={kept:<6}{flag}{broken_flag}")

    print(f"\n{tag}[6/7] Writing JSON chunks output...")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(chunks, f, ensure_ascii=False, indent=2)
    print(f"  → {output_path}")

    report = {
        "n_raw_entries":                      len(raw_entries),
        "folders_found":                       dict(folders_seen),
        "n_extracted":                          len(extracted),
        "n_excluded_broken_folder":              dict(broken_counts),
        "known_broken_folders":                   sorted(KNOWN_BROKEN_FOLDERS),
        "n_used_generic_extractor_by_folder":      dict(generic_counts),
        "folders_needing_dedicated_extractor":      sorted(FOLDERS_NEEDING_DEDICATED_EXTRACTOR),
        "n_passed_quality_filters":                  len(passed),
        "n_rejected_quality_filters":                 len(rejected),
        "rejection_reasons":                           dict(rejection_reasons),
        "rejection_reasons_by_folder":                  {k: dict(v) for k, v in rejection_reasons_by_folder.items()},
        "n_duplicates_removed":                          n_duplicates,
        "n_final_chunks":                                 len(chunks),
        "n_final_chunks_via_generic_extractor":            n_generic,
        "final_by_source":                                  dict(by_source),
        "final_by_language":                                 dict(by_language),
        "final_by_country":                                   dict(by_country),
        "per_folder_funnel": {
            folder: {
                "root":      "+".join(sorted(folder_roots.get(folder, []))),
                "found":     folders_seen.get(folder, 0),
                "extracted":  extracted_by_folder.get(folder, 0),
                "kept":        deduped_by_folder.get(folder, 0),
                "broken":       folder in KNOWN_BROKEN_FOLDERS,
            }
            for folder in sorted(folders_seen)
        },
    }
    report_path.parent.mkdir(parents=True, exist_ok=True)
    with open(report_path, "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)
    print(f"  → {report_path}")

    print(f"\n{tag}[7/7] Writing Excel report...")
    xlsx_report_path.parent.mkdir(parents=True, exist_ok=True)
    build_xlsx_report(report, xlsx_report_path)
    print(f"  → {xlsx_report_path}")

    return report


# ─── CLI: runs the pipeline 3 times (raw, raw_2, all) ─────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Check usefulness and standardize scraped wine data; produces 3 report pairs (raw / raw_2 / all)",
        formatter_class=argparse.RawTextHelpFormatter,
    )
    parser.add_argument("--input-dir", default=str(DEFAULT_INPUT_DIR))
    parser.add_argument("--input-dir-2", default=str(DEFAULT_INPUT_DIR_2))
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR))
    parser.add_argument(
        "--min-length", type=int, default=DEFAULT_MIN_LENGTH,
        help=f"Minimum text length in characters (default: {DEFAULT_MIN_LENGTH})",
    )
    args = parser.parse_args()

    input_root   = Path(args.input_dir) if args.input_dir else None
    input_root_2 = Path(args.input_dir_2) if args.input_dir_2 else None
    output_dir   = Path(args.output_dir)

    raw_paths   = sorted(input_root.rglob("*.json")) if input_root and input_root.exists() else []
    raw2_paths  = sorted(input_root_2.glob("*.json")) if input_root_2 and input_root_2.exists() else []

    if not raw_paths and not raw2_paths:
        print(
            f"ERROR: no input files found (input_dir={args.input_dir}, input_dir_2={args.input_dir_2})",
            file=sys.stderr,
        )
        sys.exit(1)

    runs = []
    if raw_paths:
        runs.append(dict(
            label="raw",
            input_paths=raw_paths,
            input_root=input_root,
            input_root_2=None,
            output=output_dir / "standardized_chunks_raw.json",
            report=output_dir / "usefulness_raw_report.json",
            xlsx=output_dir / "usefulness_raw_report.xlsx",
        ))
    if raw2_paths:
        runs.append(dict(
            label="raw_2",
            input_paths=raw2_paths,
            input_root=None,
            input_root_2=input_root_2,
            output=output_dir / "standardized_chunks_raw_2.json",
            report=output_dir / "usefulness_raw_2_report.json",
            xlsx=output_dir / "usefulness_raw_2_report.xlsx",
        ))
    if raw_paths and raw2_paths:
        runs.append(dict(
            label="all",
            input_paths=sorted(set(raw_paths + raw2_paths)),
            input_root=input_root,
            input_root_2=input_root_2,
            output=output_dir / "standardized_chunks_all.json",
            report=output_dir / "usefulness_all_report.json",
            xlsx=output_dir / "usefulness_all_report.xlsx",
        ))

    for r in runs:
        print(f"\n{'='*70}\nRUN: {r['label']}\n{'='*70}")
        run(
            input_paths=r["input_paths"],
            input_root=r["input_root"],
            output_path=r["output"],
            report_path=r["report"],
            xlsx_report_path=r["xlsx"],
            min_length=args.min_length,
            input_root_2=r["input_root_2"],
            label=r["label"],
        )


if __name__ == "__main__":
    main()